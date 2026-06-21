import asyncio
import openpyxl
from sqlalchemy import text, select
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession

from backend.app.core.config import settings
from backend.app.db.session import Base
from backend.app.db.init_db import init_public_db, init_tenant_db
from backend.app.core.security import get_password_hash
from backend.app.models.catalog import (
    Empresa, Usuario, TapaValvula, AnomaliaNeumatico,
    AnomaliaAro, DisenoReencauche, ConfiguracionVehiculo,
    UmbralRtd, UmbralPresion
)

async def seed_all():
    # 1. Connect and initialize DB schemas & tables
    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    
    print("Initializing public schema and tables...")
    async with engine.begin() as conn:
        await init_public_db(conn)
    print("Public schema initialized.")
    
    # Open the excel sheet
    wb = openpyxl.load_workbook('../REPORTES Y PATRON.xlsx', data_only=True)
    
    # 2. Seed default companies (Empresas)
    empresas_to_create = [
        {"nombre": "Palomino", "ruc": "20123456789"},
        {"nombre": "Carapongo", "ruc": "20987654321"}
    ]
    
    created_empresas = {}
    
    # Create companies in a separate block and commit
    async with AsyncSession(engine) as session:
        for emp_data in empresas_to_create:
            result = await session.execute(
                select(Empresa).where(Empresa.nombre == emp_data["nombre"])
            )
            emp = result.scalar_one_or_none()
            if not emp:
                emp = Empresa(nombre=emp_data["nombre"], ruc=emp_data["ruc"])
                session.add(emp)
                await session.flush()
            created_empresas[emp_data["nombre"]] = emp.id
        await session.commit()
            
    print(f"Companies created/found: {created_empresas}")
    
    # 3. Initialize tenant databases for each empresa (no session active, so no locks held!)
    for emp_id in created_empresas.values():
        print(f"Initializing tenant db for empresa {emp_id}...")
        async with engine.begin() as conn:
            await init_tenant_db(conn, emp_id)
        print(f"Tenant db for empresa {emp_id} initialized.")
        
    # 4. Create users, catalogs, thresholds
    async with AsyncSession(engine) as session:
        users_to_create = [
            {"nombre": "Admin Renova", "email": "admin@renova.com", "pass": "admin123", "rol": "admin", "emp_id": None},
            {"nombre": "Inspector Palomino", "email": "inspector@palomino.com", "pass": "palomino123", "rol": "inspector", "emp_id": created_empresas["Palomino"]},
            {"nombre": "Supervisor Palomino", "email": "supervisor@palomino.com", "pass": "palomino123", "rol": "supervisor", "emp_id": created_empresas["Palomino"]},
            {"nombre": "Inspector Carapongo", "email": "inspector@carapongo.com", "pass": "carapongo123", "rol": "inspector", "emp_id": created_empresas["Carapongo"]},
            {"nombre": "Supervisor Carapongo", "email": "supervisor@carapongo.com", "pass": "carapongo123", "rol": "supervisor", "emp_id": created_empresas["Carapongo"]},
        ]
        
        for u_data in users_to_create:
            result = await session.execute(
                select(Usuario).where(Usuario.email == u_data["email"])
            )
            u = result.scalar_one_or_none()
            if not u:
                u = Usuario(
                    nombre=u_data["nombre"],
                    email=u_data["email"],
                    password_hash=get_password_hash(u_data["pass"]),
                    rol=u_data["rol"],
                    empresa_id=u_data["emp_id"]
                )
                session.add(u)
        
        # Seed catalog from PATRON sheet
        ws_patron = wb['PATRON']
        
        # Tapas valvula
        tapa_valvulas = set()
        for r in range(2, ws_patron.max_row + 1):
            val = ws_patron.cell(row=r, column=1).value
            if val:
                tapa_valvulas.add(val.strip())
                
        for t in sorted(tapa_valvulas):
            result = await session.execute(
                select(TapaValvula).where(TapaValvula.nombre == t)
            )
            tv = result.scalar_one_or_none()
            if not tv:
                session.add(TapaValvula(nombre=t))
                
        # Anomalías neumático
        anomalias_neumatico = []
        for r in range(2, ws_patron.max_row + 1):
            nombre = ws_patron.cell(row=r, column=3).value
            posible_causa = ws_patron.cell(row=r, column=4).value
            desecho_str = ws_patron.cell(row=r, column=5).value
            if nombre:
                desecho = (desecho_str.upper() == 'SI' if desecho_str else False)
                anomalias_neumatico.append({
                    "nombre": nombre.strip(),
                    "posible_causa": posible_causa.strip() if posible_causa else None,
                    "desecho": desecho
                })
                
        for anom in anomalias_neumatico:
            result = await session.execute(
                select(AnomaliaNeumatico).where(AnomaliaNeumatico.nombre == anom["nombre"])
            )
            an = result.scalar_one_or_none()
            if not an:
                session.add(AnomaliaNeumatico(
                    nombre=anom["nombre"],
                    posible_causa=anom["posible_causa"],
                    desecho=anom["desecho"]
                ))
                
        # Configuraciones de Vehículos
        configs = []
        for r in range(2, ws_patron.max_row + 1):
            tipo = ws_patron.cell(row=r, column=10).value
            config = ws_patron.cell(row=r, column=11).value
            pos = ws_patron.cell(row=r, column=12).value
            eje = ws_patron.cell(row=r, column=13).value
            piso_str = ws_patron.cell(row=r, column=14).value
            if tipo and config and pos:
                piso = (piso_str.upper() == 'SI' if piso_str else True)
                configs.append({
                    "tipo_vehiculo": tipo.strip(),
                    "configuracion": config.strip(),
                    "posicion": int(pos),
                    "tipo_eje": eje.strip() if eje else "Libre",
                    "piso": piso
                })
                
        for c in configs:
            result = await session.execute(
                select(ConfiguracionVehiculo).where(
                    (ConfiguracionVehiculo.tipo_vehiculo == c["tipo_vehiculo"]) &
                    (ConfiguracionVehiculo.configuracion == c["configuracion"]) &
                    (ConfiguracionVehiculo.posicion == c["posicion"])
                )
            )
            cv = result.scalar_one_or_none()
            if not cv:
                session.add(ConfiguracionVehiculo(
                    tipo_vehiculo=c["tipo_vehiculo"],
                    configuracion=c["configuracion"],
                    posicion=c["posicion"],
                    tipo_eje=c["tipo_eje"],
                    piso=c["piso"]
                ))
                
        # Seed from REPORTE sheet
        ws_reporte = wb['REPORTE']
        headers = [c.value for c in ws_reporte[1]]
        
        medida_idx = headers.index('MEDIDA') + 1
        marca_idx = headers.index('MARCA') + 1
        dis_orig_idx = headers.index('DISEÑO ORIGINAL') + 1
        dis_act_idx = headers.index('DISEÑO ACTUAL') + 1
        anom_aro_idx = headers.index('ANOMALÍA ARO') + 1
        eje_idx = headers.index('TIPO EJE') + 1
        presion_frio_idx = headers.index('PRESIÓN ESTABLECIDO EN FRÍO') + 1
        rtd_cambio_idx = headers.index('RTD PARA CAMBIO') + 1
        rtd_proximo_idx = headers.index('RTD PRÓXIMO CAMBIO') + 1
        rtd_normal_idx = headers.index('RTD NORMAL') + 1
        
        anomalias_aro = set()
        disenos = set()
        rtd_thresholds = set()
        presion_thresholds = set()
        
        for r in range(2, ws_reporte.max_row + 1):
            medida = ws_reporte.cell(row=r, column=medida_idx).value
            marca = ws_reporte.cell(row=r, column=marca_idx).value
            dis_orig = ws_reporte.cell(row=r, column=dis_orig_idx).value
            dis_act = ws_reporte.cell(row=r, column=dis_act_idx).value
            anom_aro = ws_reporte.cell(row=r, column=anom_aro_idx).value
            eje = ws_reporte.cell(row=r, column=eje_idx).value
            presion_frio = ws_reporte.cell(row=r, column=presion_frio_idx).value
            rtd_cambio = ws_reporte.cell(row=r, column=rtd_cambio_idx).value
            rtd_proximo = ws_reporte.cell(row=r, column=rtd_proximo_idx).value
            rtd_normal = ws_reporte.cell(row=r, column=rtd_normal_idx).value
            
            if anom_aro and anom_aro.strip() != "Normal":
                anomalias_aro.add(anom_aro.strip())
            
            if medida:
                medida = medida.strip()
                if rtd_cambio is not None and rtd_proximo is not None and rtd_normal is not None:
                    rtd_thresholds.add((medida, float(rtd_cambio), float(rtd_proximo), float(rtd_normal)))
                if eje and presion_frio is not None:
                    presion_thresholds.add((medida, eje.strip(), float(presion_frio)))
                    
            if marca:
                marca = marca.strip()
                if dis_orig and dis_orig.strip():
                    disenos.add((marca, dis_orig.strip()))
                if dis_act and dis_act.strip():
                    disenos.add((marca, dis_act.strip()))
                    
        anomalias_aro.add("Normal")
        
        for aro in sorted(anomalias_aro):
            result = await session.execute(
                select(AnomaliaAro).where(AnomaliaAro.nombre == aro)
            )
            aa = result.scalar_one_or_none()
            if not aa:
                session.add(AnomaliaAro(nombre=aro))
                
        for d_marca, d_nombre in sorted(disenos):
            result = await session.execute(
                select(DisenoReencauche).where(
                    (DisenoReencauche.marca == d_marca) &
                    (DisenoReencauche.nombre == d_nombre)
                )
            )
            dr = result.scalar_one_or_none()
            if not dr:
                session.add(DisenoReencauche(marca=d_marca, nombre=d_nombre))
                
        for m, c, p, n in sorted(rtd_thresholds):
            result = await session.execute(
                select(UmbralRtd).where(
                    (UmbralRtd.medida == m) &
                    (UmbralRtd.empresa_id == None)
                )
            )
            urtd = result.scalar_one_or_none()
            if not urtd:
                session.add(UmbralRtd(medida=m, rtd_cambio=c, rtd_proximo=p, rtd_normal=n, empresa_id=None))
                
        for m, e, pf in sorted(presion_thresholds):
            result = await session.execute(
                select(UmbralPresion).where(
                    (UmbralPresion.medida == m) &
                    (UmbralPresion.tipo_eje == e) &
                    (UmbralPresion.empresa_id == None)
                )
            )
            upres = result.scalar_one_or_none()
            if not upres:
                session.add(UmbralPresion(medida=m, tipo_eje=e, presion_frio=pf, empresa_id=None))
                
        await session.commit()
        print("Data seeding completed successfully!")

if __name__ == "__main__":
    asyncio.run(seed_all())
