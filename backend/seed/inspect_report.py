import openpyxl

def main():
    wb = openpyxl.load_workbook('REPORTES Y PATRON.xlsx', data_only=True)
    ws = wb['REPORTE']

    headers = [cell.value for cell in ws[1]]
    
    medida_idx = headers.index('MEDIDA')
    eje_idx = headers.index('TIPO EJE')
    ref_idx = headers.index('PRESIÓN ESTABLECIDO EN FRÍO')
    rtd_cambio_idx = headers.index('RTD PARA CAMBIO')
    rtd_proximo_idx = headers.index('RTD PRÓXIMO CAMBIO')
    rtd_normal_idx = headers.index('RTD NORMAL')
    
    unique_pressures = set()
    unique_rtds = set()
    
    for r in range(2, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[r]]
        medida = row_vals[medida_idx]
        eje = row_vals[eje_idx]
        ref = row_vals[ref_idx]
        rtd_cambio = row_vals[rtd_cambio_idx]
        rtd_proximo = row_vals[rtd_proximo_idx]
        rtd_normal = row_vals[rtd_normal_idx]
        
        if medida is not None:
            unique_pressures.add((medida, eje, ref))
            unique_rtds.add((medida, rtd_cambio, rtd_proximo, rtd_normal))
            
    print("--- Unique Pressure References in Excel ---")
    for item in sorted(unique_pressures):
        print(f"Medida: {item[0]} | Eje: {item[1]} | Ref Cold Pressure: {item[2]}")
        
    print("\n--- Unique RTD Thresholds in Excel ---")
    for item in sorted(unique_rtds):
        print(f"Medida: {item[0]} | Cambio: {item[1]} | Próximo: {item[2]} | Normal: {item[3]}")

if __name__ == '__main__':
    main()
